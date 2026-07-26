#!/usr/bin/env python3
"""
Reusable template for generating formatted Excel checklists via openpyxl.
Usage: run inline in generate_code tool or adapt as needed.
Template by checklist-xlsx skill.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()
ws = wb.active

# === STYLES ===
header_font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
section_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
item_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
dark_gray = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

# === TITLE ===
ws.merge_cells('A1:F1')
c = ws['A1']
c.value = 'CHECKLIST TITLE'
c.font = header_font; c.fill = header_fill
c.alignment = Alignment(horizontal='center', vertical='center'); c.border = thin_border
for col in range(2, 7):
    ws.cell(row=1, column=col).fill = header_fill
    ws.cell(row=1, column=col).border = thin_border

ws.merge_cells('A2:F2')
ws['A2'].value = f'Dibuat: {datetime.now().strftime("%d/%m/%Y")} | Sel kuning = INPUT'
ws['A2'].font = Font(name='Arial', size=9, italic=True, color='000000')
ws['A2'].alignment = Alignment(horizontal='center')

# === HEADERS ===
headers = ['No', 'Item', 'Siap?', 'Keterangan', 'Catatan', 'Priority']
widths = [5, 42, 16, 32, 32, 16]
for i, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    cell.fill = dark_gray
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(i)].width = w

# === DATA: replace this list with your own ===
sections = [
    ("CATEGORY", PatternFill(start_color='C00000', end_color='C00000', fill_type='solid'), [
        ("Item Name", "Description note", "CRITICAL"),
    ]),
]

row = 4
for title, fill, items in sections:
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = title
    cell.font = section_font; cell.fill = fill
    cell.alignment = Alignment(horizontal='left', vertical='center'); cell.border = thin_border
    for c in range(2, 7):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).border = thin_border
    row += 1

    for no, (item, note, prio) in enumerate(items, 1):
        ws.cell(row=row, column=1, value=no).font = item_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

        cell_b = ws.cell(row=row, column=2, value=item)
        cell_b.font = item_font; cell_b.alignment = Alignment(vertical='center')
        cell_b.border = thin_border

        cell_c = ws.cell(row=row, column=3, value='')
        cell_c.font = Font(name='Arial', size=12, color='008000')
        cell_c.alignment = Alignment(horizontal='center')
        cell_c.fill = input_fill; cell_c.border = thin_border

        cell_d = ws.cell(row=row, column=4, value=note)
        cell_d.font = Font(name='Arial', size=9, italic=True, color='666666')
        cell_d.alignment = Alignment(vertical='center'); cell_d.border = thin_border

        cell_e = ws.cell(row=row, column=5, value='')
        cell_e.font = item_font
        cell_e.alignment = Alignment(vertical='center', wrap_text=True)
        cell_e.fill = input_fill; cell_e.border = thin_border

        cell_f = ws.cell(row=row, column=6, value=prio)
        cell_f.font = Font(name='Arial', size=9, bold=True)
        cell_f.alignment = Alignment(horizontal='center')
        if prio == 'CRITICAL':
            cell_f.font = Font(name='Arial', size=9, bold=True, color='C00000')
        elif prio == 'HIGH':
            cell_f.font = Font(name='Arial', size=9, bold=True, color='ED7D31')
        else:
            cell_f.font = Font(name='Arial', size=9, color='666666')
        cell_f.border = thin_border

        ws.row_dimensions[row].height = 22
        row += 1

# === LEGEND ===
row += 1
ws.merge_cells(f'A{row}:F{row}')
lc = ws[f'A{row}']
lc.value = 'LEGENDA'; lc.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
lc.fill = dark_gray; lc.alignment = Alignment(horizontal='center')
lc.border = thin_border
for c in range(2, 7):
    ws.cell(row=row, column=c).fill = dark_gray
    ws.cell(row=row, column=c).border = thin_border

row += 1
for leg_title, leg_desc in [
    ('Kolom kuning = INPUT', 'Diisi manual user: centang + catatan'),
    ('CRITICAL', 'Harus ada'),
    ('HIGH', 'Sangat disarankan'),
    ('NORMAL', 'Nice-to-have'),
]:
    ws.cell(row=row, column=2, value=leg_title).font = Font(name='Arial', bold=True, size=10)
    ws.cell(row=row, column=4, value=leg_desc).font = Font(name='Arial', size=9, italic=True, color='666666')
    row += 1

# === FINALIZE ===
ws.freeze_panes = 'A4'
ws.auto_filter.ref = f'A3:F{row-1}'

outpath = '/tmp/checklist_output.xlsx'
wb.save(outpath)
print(f'Saved: {outpath}')
